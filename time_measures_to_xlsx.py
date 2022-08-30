#Imports the values of the time measures to an excel file
import openpyxl

if __name__ == '__main__':
    workbook = openpyxl.load_workbook('solver_comparison.xlsx')
    sheet = workbook.active

    time_stats=open('time_measures/time_stats_solver_10_solver_25_08.txt')
    time_lines=time_stats.readlines()

    file_col=1
    dis_col=2
    input_col=3
    smt2_col=4
    ktest_col=5
    sym_var_col=6

    current_row=sheet.max_row+1
    for line in time_lines:
        values=line.split(' ')
        file_name=values[0]
        dis_time=float(values[2])
        inpt_time=float(values[4])
        smt2=int(values[6])
        ktest=int(values[8])
        sym_args=int(values[10])

        sheet.cell(row=current_row,column=file_col).value= file_name
        sheet.cell(row=current_row,column=dis_col).value= dis_time
        sheet.cell(row=current_row,column=input_col).value= inpt_time
        sheet.cell(row=current_row,column=smt2_col).value=smt2
        sheet.cell(row=current_row,column=ktest_col).value= ktest
        sheet.cell(row=current_row,column=sym_var_col).value= sym_args
         
        current_row+=1
    
    workbook.save('solver_comparison.xlsx')